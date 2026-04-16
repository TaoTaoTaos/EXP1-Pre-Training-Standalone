from __future__ import annotations

from pathlib import Path

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from reportlab.lib.units import inch

from lakeice_ncde.batch import BatchExperimentResult, collect_batch_run_snapshot, parse_batch_experiment_specs, write_batch_summary_artifacts
from lakeice_ncde.config import load_config
from lakeice_ncde.experiment.tracker import create_run_context
from lakeice_ncde.utils.io import save_dataframe, save_json, save_yaml
from lakeice_ncde.utils.paths import build_pdf_name
from lakeice_ncde.visualization.batch_pdf_report import _figure_to_reportlab_image
from lakeice_ncde.visualization.plots import create_comparison_metric_bars_figure


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def _load_experiment_config(name: str) -> dict:
    return load_config(
        project_root=PROJECT_ROOT,
        config_path=PROJECT_ROOT / "configs" / "experiments" / name,
        override_paths=[],
    )


def _create_fake_run(
    tmp_path: Path,
    experiment_name: str,
    sequence: int,
    val_rmse: float,
    test_rmse: float,
) -> BatchExperimentResult:
    run_name = f"[{sequence:02d}]_{experiment_name}_20260416_120000"
    run_dir = tmp_path / experiment_name / run_name
    artifacts_dir = run_dir / "artifacts"
    artifacts_dir.mkdir(parents=True, exist_ok=True)

    save_yaml(
        {
            "experiment": {"name": experiment_name},
            "custom_split": {"target_lake_test_start": "2026-01-01"},
            "coeffs": {"interpolation": "hermite"},
            "model": {"method": "rk4"},
            "window": {"window_days": 7},
            "features": {"target_transform": "none", "feature_columns": ["Air_Temperature_celsius"]},
            "train": {
                "device": "cpu",
                "batch_parallel": True,
                "batch_size": 16,
                "learning_rate": 2.0e-4,
                "weight_decay": 5.0e-4,
                "max_epochs": 500,
                "loss": "huber",
                "huber_delta": 0.1,
                "physics_loss": {"enabled": False, "rule": "stefan_growth_residual", "lambda_st": 0.0, "lambda_nn": 1.0, "init_kappa": 0.0, "grow_temp_threshold_celsius": -0.5},
            },
        },
        run_dir / "config_merged.yaml",
    )
    save_dataframe(
        pd.DataFrame(
            [
                {"split": "val", "loss": 0.01, "rmse": val_rmse, "mae": 0.1, "r2": 0.8, "bias": 0.0, "negative_count": 0},
                {"split": "test", "loss": 0.02, "rmse": test_rmse, "mae": 0.2, "r2": 0.7, "bias": 0.0, "negative_count": 0},
            ]
        ),
        run_dir / "metrics.csv",
    )
    save_json(
        {
            "best_epoch": 12,
            "best_val_rmse": val_rmse,
            "duration_seconds": 34.5,
            "final_val_loss": 0.01,
            "final_test_loss": 0.02,
        },
        run_dir / "run_summary.json",
    )
    save_dataframe(
        pd.DataFrame(
            [
                {"epoch": 1, "train_loss": 0.30, "val_loss": 0.20, "val_rmse": val_rmse + 0.10, "val_mae": 0.20, "val_r2": 0.20, "lr": 2.0e-4},
                {"epoch": 2, "train_loss": 0.18, "val_loss": 0.12, "val_rmse": val_rmse + 0.04, "val_mae": 0.14, "val_r2": 0.45, "lr": 2.0e-4},
                {"epoch": 3, "train_loss": 0.10, "val_loss": 0.08, "val_rmse": val_rmse, "val_mae": 0.10, "val_r2": 0.60, "lr": 2.0e-4},
            ]
        ),
        run_dir / "epoch_summary.csv",
    )
    prediction_rows = [
        {"lake_name": "Lake Xiaoxingkai", "sample_datetime": "2026-01-10 12:00:00", "y_true": 0.30, "y_pred": test_rmse + 0.10},
        {"lake_name": "Lake Xiaoxingkai", "sample_datetime": "2026-01-11 12:00:00", "y_true": 0.28, "y_pred": test_rmse + 0.08},
        {"lake_name": "Lake Xiaoxingkai", "sample_datetime": "2026-01-12 12:00:00", "y_true": 0.25, "y_pred": test_rmse + 0.06},
    ]
    save_dataframe(pd.DataFrame(prediction_rows), run_dir / "val_predictions.csv")
    save_dataframe(pd.DataFrame(prediction_rows), run_dir / "test_predictions.csv")
    seasonal_rows = [
        {"lake_name": "Lake Xiaoxingkai", "sample_datetime": "2026-01-08 12:00:00", "y_true": "", "y_pred": test_rmse + 0.12},
        {"lake_name": "Lake Xiaoxingkai", "sample_datetime": "2026-01-09 12:00:00", "y_true": "", "y_pred": test_rmse + 0.11},
        *prediction_rows,
    ]
    save_dataframe(pd.DataFrame(seasonal_rows), run_dir / "seasonal_rollout_predictions.csv")
    save_dataframe(pd.DataFrame(prediction_rows), run_dir / "seasonal_rollout_overlap_predictions.csv")
    save_json({"split_name": experiment_name, "target_lake": "Lake Xiaoxingkai"}, artifacts_dir / "run_manifest.json")
    (run_dir / build_pdf_name(run_name)).write_bytes(b"%PDF-1.4\n%fake\n")

    return BatchExperimentResult(
        experiment_name=experiment_name,
        run_dir=run_dir,
        config_path=PROJECT_ROOT / "configs" / "experiments" / f"{experiment_name}.yaml",
        override_paths=(),
        set_values=(),
    )


def test_run_all_config_resolves_expected_child_experiments() -> None:
    config = _load_experiment_config("Run-ALL.yaml")

    specs = parse_batch_experiment_specs(config, PROJECT_ROOT)

    assert [spec.experiment_name for spec in specs] == [
        "EXP0_pretrain_autoreg",
        "EXP1_transfer_autoreg",
        "EXP2_transfer_autoreg_stefan",
    ]
    assert specs[0].config_path.name == "EXP0_pretrain_autoreg.yaml"
    assert specs[1].config_path.name == "EXP1_transfer_autoreg.yaml"
    assert specs[2].config_path.name == "EXP2_transfer_autoreg_stefan.yaml"


def test_write_batch_summary_artifacts_copies_pdfs_and_writes_excel(tmp_path) -> None:
    result_a = _create_fake_run(tmp_path, "EXP0_pretrain_autoreg", sequence=1, val_rmse=0.11, test_rmse=0.21)
    result_b = _create_fake_run(tmp_path, "EXP1_transfer_autoreg", sequence=2, val_rmse=0.12, test_rmse=0.22)

    snapshots = [
        collect_batch_run_snapshot(result_a),
        collect_batch_run_snapshot(result_b),
    ]
    run_context = create_run_context(tmp_path / "outputs", "Run-ALL", {"experiment": {"name": "Run-ALL"}})

    excel_path, copied_pdf_paths, summary_pdf_path = write_batch_summary_artifacts(run_context, snapshots)

    assert excel_path.exists()
    assert summary_pdf_path.exists()
    assert summary_pdf_path.name.startswith("01_Run-ALL_")
    assert summary_pdf_path.suffix == ".pdf"
    assert excel_path.name.startswith("01_Run-ALL_")
    assert excel_path.name.endswith("_summary.xlsx")
    assert [path.name for path in copied_pdf_paths] == [
        "01_EXP0_pretrain_autoreg_20260416_120000.pdf",
        "02_EXP1_transfer_autoreg_20260416_120000.pdf",
    ]
    assert all(path.exists() for path in copied_pdf_paths)

    workbook = load_workbook(excel_path)
    worksheet = workbook["summary"]
    headers = [worksheet.cell(row=1, column=index).value for index in range(1, 4)]
    assert headers == ["record", "EXP0_pretrain_autoreg", "EXP1_transfer_autoreg"]

    first_column = {
        worksheet.cell(row=row_index, column=1).value: worksheet.cell(row=row_index, column=2).value
        for row_index in range(2, worksheet.max_row + 1)
    }
    assert first_column["run.run_name"] == "[01]_EXP0_pretrain_autoreg_20260416_120000"
    assert first_column["metrics.val.rmse"] == 0.11
    assert first_column["summary.best_epoch"] == 12


def test_figure_to_reportlab_image_uses_rendered_image_aspect_ratio() -> None:
    fig, ax = plt.subplots(figsize=(6, 2))
    ax.bar(range(3), [0.08, 0.14, 0.12], color=["#1f77b4", "#d62728", "#2ca02c"])
    ax.set_xticks(range(3))
    ax.set_xticklabels(
        [
            "EXP0_pretrain_autoreg",
            "EXP1_transfer_autoreg",
            "EXP2_transfer_autoreg_stefan",
        ],
        rotation=28,
        ha="right",
    )
    ax.set_ylabel("Validation RMSE")
    ax.set_title("Validation Metrics Comparison")

    image = _figure_to_reportlab_image(fig, max_width=4.0 * inch, max_height=2.5 * inch)

    original_aspect = 6.0 / 2.0
    rendered_aspect = image.imageWidth / image.imageHeight
    draw_aspect = image.drawWidth / image.drawHeight

    assert abs(rendered_aspect - original_aspect) > 0.2
    assert abs(draw_aspect - rendered_aspect) < 0.05
    assert image.drawWidth <= 4.0 * inch
    assert image.drawHeight <= 2.5 * inch


def test_comparison_metric_bars_keep_small_value_labels_close_to_bars() -> None:
    metric_table = pd.DataFrame(
        [
            {"experiment_name": "EXP0_pretrain_autoreg", "split": "val", "loss": 0.010, "rmse": 0.080, "mae": 0.060, "r2": 0.70, "bias": 0.010, "negative_count": 0},
            {"experiment_name": "EXP1_transfer_autoreg", "split": "val", "loss": 0.011, "rmse": 0.082, "mae": 0.061, "r2": 0.80, "bias": -0.010, "negative_count": 0},
            {"experiment_name": "EXP2_transfer_autoreg_stefan", "split": "val", "loss": 0.012, "rmse": 0.085, "mae": 0.062, "r2": 0.79, "bias": -0.010, "negative_count": 0},
        ]
    )

    fig = create_comparison_metric_bars_figure(
        metric_table,
        split="val",
        experiment_names=[
            "EXP0_pretrain_autoreg",
            "EXP1_transfer_autoreg",
            "EXP2_transfer_autoreg_stefan",
        ],
    )
    loss_axis = fig.axes[0]
    label_positions = [text.get_position()[1] for text in loss_axis.texts]
    plt.close(fig)

    assert label_positions
    assert max(label_positions) < 0.02
