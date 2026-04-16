from __future__ import annotations

import shutil
from collections import Counter
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(r"S:\STU-Papers\My_Papers\EXP1-Pre-Training-Standalone")
TARGET_XLSX = PROJECT_ROOT / "data" / "raw" / "LakeIce_ERA5_less60_XiaoxingkaiDaily12.xlsx"
SOURCE_CSV = Path(r"D:\浏览器下载\era5_pending_observations_refilled (2).csv")
SHEET_NAME = "lakeice_era5"
SKIP_LAKES = {"【Li】Lake Xiaoxingkai"}
UPDATE_COLUMNS = [
    "era5_datetime",
    "Ten_Meter_Elevation_Wind_Speed_meterPerSecond",
    "Air_Temperature_celsius",
    "Relative_Humidity_percent",
    "Shortwave_Radiation_Downwelling_wattPerMeterSquared",
    "Longwave_Radiation_Downwelling_wattPerMeterSquared",
    "Sea_Level_Barometric_Pressure_pascal",
    "Surface_Level_Barometric_Pressure_pascal",
    "Precipitation_millimeterPerDay",
    "Snowfall_millimeterPerDay",
]


def normalize_datetime(value) -> datetime:
    if pd.isna(value):
        raise ValueError("Encountered missing datetime value.")
    timestamp = pd.Timestamp(value)
    if timestamp.tzinfo is not None:
        timestamp = timestamp.tz_convert(None)
    return timestamp.to_pydatetime().replace(tzinfo=None)


def values_equal(left, right) -> bool:
    if pd.isna(left) and pd.isna(right):
        return True
    if isinstance(left, datetime) or isinstance(right, datetime):
        return normalize_datetime(left) == normalize_datetime(right)
    if isinstance(left, float) or isinstance(right, float):
        return abs(float(left) - float(right)) <= 1e-12
    return left == right


def main() -> None:
    if not TARGET_XLSX.exists():
        raise FileNotFoundError(f"Target workbook not found: {TARGET_XLSX}")
    if not SOURCE_CSV.exists():
        raise FileNotFoundError(f"Source CSV not found: {SOURCE_CSV}")

    source_df = pd.read_csv(SOURCE_CSV)
    source_df = source_df.loc[~source_df["lake_name"].isin(SKIP_LAKES)].copy()
    source_df["sample_datetime"] = pd.to_datetime(source_df["sample_datetime"])
    source_df["era5_datetime"] = pd.to_datetime(source_df["era5_datetime"])

    duplicate_mask = source_df.duplicated(["lake_name", "sample_datetime"], keep=False)
    if duplicate_mask.any():
        duplicates = source_df.loc[duplicate_mask, ["lake_name", "sample_datetime"]]
        raise ValueError(
            "Source CSV contains duplicate lake_name + sample_datetime keys:\n"
            f"{duplicates.head(20).to_string(index=False)}"
        )

    lookup = {
        (str(row["lake_name"]), normalize_datetime(row["sample_datetime"])): {
            column: row[column] for column in UPDATE_COLUMNS
        }
        for _, row in source_df.iterrows()
    }

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = TARGET_XLSX.with_name(f"{TARGET_XLSX.stem}.backup_{timestamp}{TARGET_XLSX.suffix}")
    shutil.copy2(TARGET_XLSX, backup_path)

    workbook = load_workbook(TARGET_XLSX)
    worksheet = workbook[SHEET_NAME]
    header_map = {worksheet.cell(1, column).value: column for column in range(1, worksheet.max_column + 1)}

    required_headers = {"lake_name", "sample_datetime", *UPDATE_COLUMNS}
    missing_headers = sorted(header for header in required_headers if header not in header_map)
    if missing_headers:
        raise ValueError(f"Workbook is missing required columns: {missing_headers}")

    updated_rows = 0
    skipped_rows = 0
    missing_keys: list[tuple[str, datetime]] = []
    changed_cells = Counter()

    for row_idx in range(2, worksheet.max_row + 1):
        lake_name = worksheet.cell(row_idx, header_map["lake_name"]).value
        if lake_name in SKIP_LAKES:
            skipped_rows += 1
            continue

        sample_datetime = worksheet.cell(row_idx, header_map["sample_datetime"]).value
        key = (str(lake_name), normalize_datetime(sample_datetime))
        source_row = lookup.get(key)
        if source_row is None:
            missing_keys.append(key)
            continue

        updated_rows += 1
        for column_name in UPDATE_COLUMNS:
            cell = worksheet.cell(row_idx, header_map[column_name])
            new_value = source_row[column_name]
            if isinstance(new_value, pd.Timestamp):
                new_value = new_value.to_pydatetime().replace(tzinfo=None)
            elif pd.isna(new_value):
                new_value = None

            if not values_equal(cell.value, new_value):
                changed_cells[column_name] += 1
                cell.value = new_value

    if missing_keys:
        preview = "\n".join(f"{lake} | {sample_dt}" for lake, sample_dt in missing_keys[:20])
        raise ValueError(f"Workbook rows missing in source CSV ({len(missing_keys)} rows):\n{preview}")

    workbook.save(TARGET_XLSX)

    print(f"Backup created: {backup_path}")
    print(f"Workbook updated: {TARGET_XLSX}")
    print(f"Updated rows: {updated_rows}")
    print(f"Skipped rows: {skipped_rows}")
    print("Changed cells by column:")
    for column_name in UPDATE_COLUMNS:
        print(f"  {column_name}: {changed_cells[column_name]}")


if __name__ == "__main__":
    main()
